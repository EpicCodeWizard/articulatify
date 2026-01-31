"""Copyright (c) 2026 Sarvesh Madullapalli (@EpicCodeWizard)"""

import json
import os
import re
import tempfile
from collections import defaultdict

import gradio as gr
import openpyxl
import requests
from openpyxl.styles import Font
from titlecase import titlecase

all_institutions = []
destination_institutions = []
academic_years = []
source_institution = {}


def flat_map(func, items):
    result = []
    for item in items:
        result.extend(func(item))
    return result


def shorten_institution_name(name):
    return name.replace("University of California, ", "UC ").replace(" Los Angeles", "LA").replace("California State University, ", "CSU ")


def normalize_case(text):
    if text == text.upper() or " AND " in text or " OR " in text:
        return titlecase(text)
    return text


def course_sort_key(course_string):
    return [course_string.split(" ")[0]] + [
        x for x in sum([
            [int(n), l]
            for part in course_string.replace("C100", "").split('+')
            for match in [re.findall(r'\d+|[A-Z]+', part.split()[-1])]
            for n, l in [(match[0], match[1] if len(match) > 1 else 'A')]
        ], [])
    ]


def append_log(logs, message):
    logs.append(message)
    return "\n".join(logs)


def format_receiving_course(course):
    result = ""
    if course["type"] == "Course":
        result = "[" + course["course"]["prefix"] + " " + course["course"]["courseNumber"] + ": " + normalize_case(course["course"]["courseTitle"].strip()) + "]"
    elif course["type"] == "Series":
        for articulation in course["series"]["courses"][:-1]:
            result += "[" + articulation["prefix"] + " " + articulation["courseNumber"] + ": " + normalize_case(articulation["courseTitle"].strip()) + "] " + course["series"]["conjunction"].upper() + " "
        articulation = course["series"]["courses"][-1]
        result += "[" + articulation["prefix"] + " " + articulation["courseNumber"] + ": " + normalize_case(articulation["courseTitle"].strip()) + "]"
    elif course["type"] == "Requirement":
        result = "[REQUIREMENT: " + course["requirement"]["name"] + "]"
    elif course["type"] == "GeneralEducation":
        result = "[GEN ED REQUIREMENT: " + course["generalEducationArea"]["name"] + "]"
    elif course["type"] == "Transferability":
        result = "[TRANSFERABLE]"
    return result


def process_articulations(articulations):
    course_mapping = {}

    for course in articulations:
        if not course["sendingArticulation"] or len(course["sendingArticulation"]["items"]) == 0:
            continue

        receiving_course = format_receiving_course(course)
        if len(receiving_course) == 0:
            raise ValueError("Unknown course type: " + str(course))

        for sending_articulation in course["sendingArticulation"]["items"]:
            if sending_articulation["type"] == "Advisement":
                continue
            sending_course = ""
            for articulation in sending_articulation["items"][:-1]:
                sending_course += "[" + articulation["prefix"] + " " + articulation["courseNumber"] + ": " + normalize_case(articulation["courseTitle"].strip()) + "] " + sending_articulation["courseConjunction"].upper() + " "
            articulation = sending_articulation["items"][-1]
            sending_course += "[" + articulation["prefix"] + " " + articulation["courseNumber"] + ": " + normalize_case(articulation["courseTitle"].strip()) + "]"
            course_mapping[sending_course] = course_mapping.get(sending_course, []) + [receiving_course]

    processed = {}
    for key, value in course_mapping.items():
        filtered = [x for x in list(set(value)) if "REQUIREMENT: " not in x and "TRANSFERABLE" not in x]
        split_courses = flat_map(lambda x: x.split(" OR "), filtered)
        if len(split_courses) > 0:
            for k in key.split(" OR "):
                to_add = processed.get(k, []) + split_courses
                processed[k] = [s for s in to_add if not any(s in other and s != other for other in to_add)]

    return processed


def process_articulation(source_school, destination_schools, year_range, include_classname, include_destination_coursename):
    logs = []

    yield None, append_log(logs, "[info] starting articulations scraping for " + source_school)

    destination_year = next(x for x in academic_years if str(x["FallYear"]) == year_range.split("-")[0])
    articulation_results = {}
    class_names = {}
    destination_course_names = {}

    yield None, append_log(logs, "[info] processing " + str(len(destination_schools)) + " destination school(s) for " + year_range + " academic year")

    for destination_school in destination_schools:
        yield None, append_log(logs, "[debug] fetching articulation data for " + destination_school)
        destination_inst = next(x for x in destination_institutions if x["institutionName"] == destination_school)

        response = requests.get("https://assist.org/api/articulation/Agreements?Key=" + str(destination_year["Id"]) + "/" + str(source_institution["id"]) + "/to/" + str(destination_inst["institutionParentId"]) + "/AllMajors").json()

        if not response.get("result"):
            yield None, append_log(logs, "[debug] trying reverse lookup for " + destination_school)
            response = requests.get("https://assist.org/api/articulation/Agreements?Key=" + str(destination_year["Id"]) + "/" + str(destination_inst["institutionParentId"]) + "/to/" + str(source_institution["id"]) + "/AllMajors").json()
            if not response.get("result"):
                yield None, append_log(logs, "[error] articulation not found for " + destination_school)
                continue

        with open("out.json", "w") as file:
            file.write(json.dumps(response, indent=2))

        raw_articulations = [x["articulation"] for x in json.loads(response["result"]["articulations"])]
        processed = process_articulations(raw_articulations)

        prefix_replacements = {}
        for course_key in processed:
            for course_part in course_key.split(" AND "):
                name = course_part[1:].split(": ")[0]
                prefix_replacements[" ".join(name.split(" ")[:-1])] = "".join(x for x in "".join(name.split(" ")[:-1]) if x.isalnum())

        destination_course_names[destination_school] = {}
        for destination_courses in processed.values():
            for destination_course in destination_courses:
                for course_segment in destination_course.split(" AND "):
                    course_id = course_segment[1:].split(": ")[0]
                    for k, v in prefix_replacements.items():
                        course_id = course_id.replace(k, v)
                    destination_course_names[destination_school][course_id] = ": ".join(course_segment[1:-1].split(": ")[1:])

        articulation_results[destination_school] = []
        for course in processed:
            course_to_add = " + ".join(sorted([x[1:].split(": ")[0] for x in course.split(" AND ")]))
            for course_segment in course.split(" AND "):
                course_id = course_segment[1:].split(": ")[0]
                for k, v in prefix_replacements.items():
                    course_id = course_id.replace(k, v)
                class_names[course_id] = ": ".join(course_segment[1:-1].split(": ")[1:])
            for k, v in prefix_replacements.items():
                course_to_add = course_to_add.replace(k, v)
            articulation_results[destination_school].append(course_to_add)
            destination_course_names[destination_school][course_to_add] = " / ".join([
                "[" + " + ".join(sorted([x[1:].split(": ")[0] + "]" for x in dc.split(" AND ")]))
                for dc in processed[course]
            ])

        articulation_results[destination_school] = list(sorted(list(set(articulation_results[destination_school])), key=course_sort_key))
        yield None, append_log(logs, "[info] completed articulations for " + destination_school + ", found " + str(len(articulation_results[destination_school])) + " courses")

    yield from generate_spreadsheet(logs, articulation_results, destination_course_names, class_names, source_school, include_classname, include_destination_coursename)


def generate_spreadsheet(logs, articulation_results, destination_course_names, class_names, source_school, include_classname, include_destination_coursename):
    yield None, append_log(logs, "[info] generating excel spreadsheet")

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    intro_sheet = workbook.create_sheet("\u200b", 0)
    intro_sheet['A1'] = "Navigate to your desired department using the spreadsheet tabs at the bottom to view university course articulations."
    intro_sheet['A1'].font = Font(size=18)
    intro_sheet['A2'] = "Generated by Articulatify"
    intro_sheet['A2'].font = Font(size=10)

    department_courses = defaultdict(set)
    for college, courses in articulation_results.items():
        for course in courses:
            for section in course.split(" + "):
                department = section.split(" ")[0]
                department_courses[department].add(course)

    colleges = list(articulation_results.keys())
    yield None, append_log(logs, "[info] creating " + str(len(department_courses)) + " department sheet(s)")

    for department, courses in sorted(department_courses.items()):
        worksheet = workbook.create_sheet(department)
        courses_sorted = list(sorted(list(set(courses)), key=course_sort_key))

        for col, course in enumerate(courses_sorted, 2):
            worksheet.cell(1, col, course)

        total_row = len(colleges) + 2
        for row, college in enumerate(colleges, 2):
            worksheet.cell(row, 1, shorten_institution_name(college))
            worksheet.cell(row + len(colleges) + 3 + (1 if include_classname else 0), 1, shorten_institution_name(college))
            for col, course in enumerate(courses_sorted, 2):
                if course in articulation_results[college]:
                    worksheet.cell(row, col, destination_course_names[college][course] if include_destination_coursename else "✔")
                if include_classname:
                    worksheet.cell(total_row + 1, col, " + ".join([class_names[x] for x in course.split(" + ")]))

        worksheet.cell(total_row, 1, "Total")
        if include_classname:
            worksheet.cell(total_row + 1, 1, "Course Name")
        worksheet.cell(total_row + len(colleges) + 3 + (1 if include_classname else 0), 1, "Total")
        if include_classname:
            worksheet.cell(total_row + len(colleges) + 5, 1, "Course Name")

        for col in range(2, len(courses_sorted) + 2):
            count_function = "COUNTA" if include_destination_coursename else "COUNTIF"
            count_arguments = "" if include_destination_coursename else ", \"✔\""
            column_letter = openpyxl.utils.get_column_letter(col)
            worksheet.cell(total_row, col, "=" + count_function + "(" + column_letter + "2:" + column_letter + str(len(colleges) + 1) + count_arguments + ")")
            worksheet.column_dimensions[column_letter].width = 14

        if include_classname:
            total_row += 1
        last_column = openpyxl.utils.get_column_letter(len(courses_sorted) + 1)
        worksheet.cell(total_row + 2, 2, "=TRANSPOSE(SORT(TRANSPOSE({B1:" + last_column + "1; B2:" + last_column + str(total_row) + "}), ROWS(B1:" + last_column + "1)+MATCH(\"Total\", A2:A" + str(total_row) + ", 0), FALSE))")
        worksheet.column_dimensions['A'].width = 20

    filename = shorten_institution_name(source_school) + " Articulations.xlsx"
    yield None, append_log(logs, "[info] saving spreadsheet as " + filename)

    temporary_directory = tempfile.mkdtemp()
    temporary_path = os.path.join(temporary_directory, filename)
    workbook.save(temporary_path)

    yield temporary_path, append_log(logs, "[info] finished articulations scraping for " + source_school)


def fetch_source_schools():
    global all_institutions
    all_institutions = requests.get("https://assist.org/api/institutions").json()
    return sorted(list(set([x["names"][0]["name"] for x in all_institutions])))


def fetch_destination_schools(source_school):
    if len(source_school) == 0:
        return gr.Dropdown(label="Destination Institutions", multiselect=True, visible=False)
    global destination_institutions, source_institution
    source_institution = next(x for x in all_institutions if x["names"][0]["name"] == source_school)
    destination_institutions = requests.get("https://assist.org/api/institutions/" + str(source_institution["id"]) + "/agreements").json()
    names = [x["institutionName"] for x in destination_institutions]
    return gr.Dropdown(choices=sorted(list(set(names))), multiselect=True, label="Destination Institutions", value=[])


def fetch_available_years(destination_schools):
    if len(destination_schools) == 0:
        return gr.Dropdown(label="Year", visible=False)
    global academic_years
    year_lists = []
    for school in destination_schools:
        destination_inst = next(x for x in destination_institutions if x["institutionName"] == school)
        academic_years = requests.get("https://assist.org/api/AcademicYears").json()
        year_lists.append([
            str(x["FallYear"]) + "-" + str(x["FallYear"] + 1)
            for x in academic_years
            if x["Id"] in destination_inst["receivingYearIds"] + destination_inst["sendingYearIds"]
        ])
    return gr.Dropdown(choices=sorted(set.intersection(*map(set, year_lists)), reverse=True), label="Year")


def show_destination_dropdown():
    return gr.Dropdown(visible=True)


def show_year_dropdown():
    return gr.Dropdown(visible=True)


with gr.Blocks(title="Articulatify") as interface:
    gr.Markdown("# Articulatify")
    gr.Markdown("### Generate bulk course articulations spreadsheets using Assist.org.")

    source = gr.Dropdown(choices=fetch_source_schools(), label="Source Institution", value=None)
    destination = gr.Dropdown(label="Destination Institutions", multiselect=True, visible=False)
    year = gr.Dropdown(label="Year", visible=False)
    include_classname = gr.Checkbox(label="Include Class Names", value=False)
    include_destination_coursename = gr.Checkbox(label="Include Destination Course Number", value=True)

    with gr.Row():
        submit = gr.Button("Generate")

    output = gr.File(label="Download Spreadsheet")
    logs = gr.Textbox(label="Logs", lines=10, max_lines=20, interactive=False, autoscroll=False, elem_id="logs-box")

    source.change(
        fn=show_destination_dropdown,
        outputs=destination
    ).then(
        fn=fetch_destination_schools,
        inputs=source,
        outputs=destination
    )

    destination.change(
        fn=show_year_dropdown,
        outputs=year
    ).then(
        fn=fetch_available_years,
        inputs=destination,
        outputs=year
    )

    submit.click(
        fn=process_articulation,
        inputs=[source, destination, year, include_classname, include_destination_coursename],
        outputs=[output, logs]
    )

if __name__ == "__main__":
    interface.launch(
        pwa=True,
        css="footer{display:none !important}\n#logs-box .generating {border:none !important}\n#logs-box textarea {scroll-behavior:auto !important}\n#logs-box.generating {border-color:var(--border-color-primary) !important}\n.generating {border-color:var(--border-color-primary) !important}",
        favicon_path="icon.png",
        server_name="0.0.0.0"
    )
